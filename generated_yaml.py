import os
import sys

from openpyxl import load_workbook
import yaml


def read_import_temple():
    """
    格式化表头内容存入list
    :return:
    """
    file_list = os.listdir()
    try:
        for file in file_list:
            if '.xlsx' in file:
                file_name = file
        wb = load_workbook(file_name)
    except:
        print('请将导入模板文件放置当前目录下再次运行')
        sys.exit()
    ws = wb['Sheet1']
    header_list = list()
    if ws['A1'].comment is not None:
        header_row = ws[1]
    else:
        header_row = ws[2]
    for row in header_row:
        header_list.append(row.value)
    for i in range(len(header_list)):
        header_list[i] = {'name': header_list[i], 'content': None, 'length': None}
    return header_list


def append_dicts_to_yaml_file(data_list, yaml_file, Dumper=yaml.SafeDumper):
    """
    将字典列表逐一追加到YAML文件中
    :param Dumper:
    :param data_list: 包含字典的列表
    :param yaml_file: 要写入的YAML文件路径
    """

    class OrderedDumper(Dumper):
        pass

    # 这里是 把 生成文件里的 “null” 转为 “”
    def represent_none(self, _):
        return self.represent_scalar('tag:yaml.org,2002:null', '')

    # 将所有数据写入YAML文件
    with open(yaml_file, 'w', encoding='utf-8') as f:
        OrderedDumper.add_representer(type(None), represent_none)
        yaml.dump(data_list, f, OrderedDumper, default_flow_style=False, indent=2, allow_unicode=True, sort_keys=False)

    print('yaml文件更新成功')


if __name__ == '__main__':
    res = read_import_temple()
    append_dicts_to_yaml_file(res, 'data_rules.yaml')
